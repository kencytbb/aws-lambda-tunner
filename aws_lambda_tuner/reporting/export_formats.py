"""
Extended export formats for AWS Lambda Tuner reports.
Provides PDF, Excel, and other advanced export capabilities.
"""

import io
import base64
from typing import Dict, Any, List, Optional, Union
from pathlib import Path
import logging
from datetime import datetime
import json

# PDF generation
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.platypus import PageBreak, KeepTogether
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.lib.colors import HexColor

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

    # Create mock classes for type hints when reportlab is not available
    class Drawing:
        pass


# Excel generation
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

    # Create mock classes for type hints when pandas/openpyxl is not available
    class pd:
        class ExcelWriter:
            pass

    class Workbook:
        pass


from ..utils import calculate_cost, calculate_statistics
from ..exceptions import ReportGenerationError

logger = logging.getLogger(__name__)


class PDFExporter:
    """Exports Lambda tuning reports to PDF format."""

    def __init__(self, results: Dict[str, Any]):
        """
        Initialize PDF exporter.

        Args:
            results: Tuning results dictionary
        """
        if not REPORTLAB_AVAILABLE:
            raise ReportGenerationError(
                "ReportLab is required for PDF export. Install with: pip install reportlab"
            )

        self.results = results
        self.configurations = results.get("configurations", [])
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()

    def _setup_custom_styles(self):
        """Setup custom styles for the PDF."""
        # Title style
        self.styles.add(
            ParagraphStyle(
                name="CustomTitle",
                parent=self.styles["Title"],
                fontSize=24,
                spaceAfter=30,
                textColor=colors.HexColor("#2c3e50"),
                alignment=TA_CENTER,
            )
        )

        # Section header style
        self.styles.add(
            ParagraphStyle(
                name="SectionHeader",
                parent=self.styles["Heading1"],
                fontSize=16,
                spaceAfter=12,
                spaceBefore=20,
                textColor=colors.HexColor("#34495e"),
                borderWidth=1,
                borderColor=colors.HexColor("#3498db"),
                borderPadding=5,
                backColor=colors.HexColor("#ecf0f1"),
            )
        )

        # Metric style
        self.styles.add(
            ParagraphStyle(
                name="MetricValue",
                parent=self.styles["Normal"],
                fontSize=14,
                alignment=TA_CENTER,
                textColor=colors.HexColor("#27ae60"),
                fontName="Helvetica-Bold",
            )
        )

    def export_performance_report(self, output_path: str, workload_type: str = "generic"):
        """
        Export a comprehensive performance report to PDF.

        Args:
            output_path: Path to save the PDF
            workload_type: Type of workload for specialized reporting
        """
        try:
            doc = SimpleDocTemplate(
                output_path,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18,
            )

            story = []

            # Title page
            story.extend(self._create_title_page(workload_type))
            story.append(PageBreak())

            # Executive summary
            story.extend(self._create_executive_summary())
            story.append(PageBreak())

            # Detailed analysis
            story.extend(self._create_detailed_analysis())
            story.append(PageBreak())

            # Charts and visualizations
            story.extend(self._create_charts_section())
            story.append(PageBreak())

            # Recommendations
            story.extend(self._create_recommendations_section())

            # Build PDF
            doc.build(story)
            logger.info(f"PDF report exported to {output_path}")

        except Exception as e:
            logger.error(f"Error exporting PDF report: {e}")
            raise ReportGenerationError(f"Failed to export PDF report: {e}")

    def _create_title_page(self, workload_type: str) -> List:
        """Create title page elements."""
        elements = []

        # Main title
        title = Paragraph(f"AWS Lambda Performance Analysis", self.styles["CustomTitle"])
        elements.append(title)
        elements.append(Spacer(1, 20))

        # Workload type
        workload_title = Paragraph(
            f"Workload Type: {workload_type.title()}", self.styles["Heading2"]
        )
        elements.append(workload_title)
        elements.append(Spacer(1, 30))

        # Summary metrics
        optimal_config = self._find_optimal_configuration()
        if optimal_config:
            summary_data = [
                ["Metric", "Value"],
                ["Optimal Memory Configuration", f"{optimal_config['memory_mb']}MB"],
                ["Average Duration", f"{optimal_config['avg_duration']:.2f}ms"],
                ["Cost per Invocation", f"${optimal_config['avg_cost']:.6f}"],
                ["Total Configurations Tested", str(len(self.configurations))],
                ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ]

            summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
            summary_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3498db")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 12),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            )

            elements.append(summary_table)

        return elements

    def _create_executive_summary(self) -> List:
        """Create executive summary section."""
        elements = []

        elements.append(Paragraph("Executive Summary", self.styles["SectionHeader"]))
        elements.append(Spacer(1, 12))

        # Find optimal configuration
        optimal = self._find_optimal_configuration()
        baseline = self._get_baseline_stats()

        if optimal and baseline:
            performance_gain = (
                (
                    (baseline["avg_duration"] - optimal["avg_duration"])
                    / baseline["avg_duration"]
                    * 100
                )
                if baseline["avg_duration"] > 0
                else 0
            )
            cost_savings = (
                ((baseline["avg_cost"] - optimal["avg_cost"]) / baseline["avg_cost"] * 100)
                if baseline["avg_cost"] > 0
                else 0
            )

            summary_text = f"""
            <b>Performance Optimization Results:</b><br/>
            <br/>
            • Optimal memory configuration: <b>{optimal['memory_mb']}MB</b><br/>
            • Performance improvement: <b>{performance_gain:.1f}%</b><br/>
            • Cost savings: <b>{cost_savings:.1f}%</b><br/>
            • Average execution time: <b>{optimal['avg_duration']:.2f}ms</b><br/>
            • Cost per invocation: <b>${optimal['avg_cost']:.6f}</b><br/>
            <br/>
            <b>Key Findings:</b><br/>
            • Tested {len(self.configurations)} different memory configurations<br/>
            • Total executions analyzed: {self._count_total_executions()}<br/>
            • Optimization strategy provides significant performance and cost benefits<br/>
            """

            elements.append(Paragraph(summary_text, self.styles["Normal"]))

        return elements

    def _create_detailed_analysis(self) -> List:
        """Create detailed analysis section."""
        elements = []

        elements.append(Paragraph("Detailed Performance Analysis", self.styles["SectionHeader"]))
        elements.append(Spacer(1, 12))

        # Configuration comparison table
        table_data = [
            [
                "Memory (MB)",
                "Avg Duration (ms)",
                "Min (ms)",
                "Max (ms)",
                "Cost per Invocation ($)",
                "Success Rate (%)",
            ]
        ]

        for config in self.configurations:
            memory_mb = config["memory_mb"]
            executions = config.get("executions", [])
            successful = [e for e in executions if not e.get("error")]

            if successful:
                durations = [e["duration"] for e in successful]
                costs = [
                    calculate_cost(memory_mb, e.get("billed_duration", e["duration"]))
                    for e in successful
                ]

                stats = calculate_statistics(durations)
                avg_cost = sum(costs) / len(costs)
                success_rate = len(successful) / len(executions) * 100

                table_data.append(
                    [
                        str(memory_mb),
                        f"{stats['mean']:.2f}",
                        f"{stats['min']:.2f}",
                        f"{stats['max']:.2f}",
                        f"{avg_cost:.6f}",
                        f"{success_rate:.1f}",
                    ]
                )

        if len(table_data) > 1:
            config_table = Table(
                table_data,
                colWidths=[1 * inch, 1.2 * inch, 0.8 * inch, 0.8 * inch, 1.5 * inch, 1 * inch],
            )
            config_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("FONTSIZE", (0, 1), (-1, -1), 9),
                    ]
                )
            )

            elements.append(config_table)

        return elements

    def _create_charts_section(self) -> List:
        """Create charts and visualizations section."""
        elements = []

        elements.append(Paragraph("Performance Visualizations", self.styles["SectionHeader"]))
        elements.append(Spacer(1, 12))

        # Performance chart
        performance_chart = self._create_performance_chart()
        if performance_chart:
            elements.append(performance_chart)
            elements.append(Spacer(1, 20))

        # Cost chart
        cost_chart = self._create_cost_chart()
        if cost_chart:
            elements.append(cost_chart)

        return elements

    def _create_performance_chart(self) -> Optional[Drawing]:
        """Create performance bar chart."""
        try:
            drawing = Drawing(400, 300)
            chart = VerticalBarChart()
            chart.x = 50
            chart.y = 50
            chart.height = 200
            chart.width = 300

            # Extract data
            memory_sizes = []
            avg_durations = []

            for config in self.configurations:
                memory_sizes.append(config["memory_mb"])
                executions = config.get("executions", [])
                successful = [e for e in executions if not e.get("error")]

                if successful:
                    durations = [e["duration"] for e in successful]
                    avg_durations.append(sum(durations) / len(durations))
                else:
                    avg_durations.append(0)

            if avg_durations:
                chart.data = [avg_durations]
                chart.categoryAxis.categoryNames = [f"{m}MB" for m in memory_sizes]
                chart.valueAxis.valueMin = 0
                chart.valueAxis.valueMax = max(avg_durations) * 1.1

                chart.bars[0].fillColor = colors.HexColor("#3498db")
                chart.categoryAxis.labels.boxAnchor = "n"
                chart.categoryAxis.labels.angle = 45

                drawing.add(chart)
                return drawing

        except Exception as e:
            logger.warning(f"Failed to create performance chart: {e}")

        return None

    def _create_cost_chart(self) -> Optional[Drawing]:
        """Create cost analysis chart."""
        try:
            drawing = Drawing(400, 300)
            chart = VerticalBarChart()
            chart.x = 50
            chart.y = 50
            chart.height = 200
            chart.width = 300

            # Extract cost data
            memory_sizes = []
            avg_costs = []

            for config in self.configurations:
                memory_mb = config["memory_mb"]
                memory_sizes.append(memory_mb)
                executions = config.get("executions", [])
                successful = [e for e in executions if not e.get("error")]

                if successful:
                    costs = [
                        calculate_cost(memory_mb, e.get("billed_duration", e["duration"]))
                        for e in successful
                    ]
                    avg_costs.append(sum(costs) / len(costs))
                else:
                    avg_costs.append(0)

            if avg_costs:
                chart.data = [avg_costs]
                chart.categoryAxis.categoryNames = [f"{m}MB" for m in memory_sizes]
                chart.valueAxis.valueMin = 0
                chart.valueAxis.valueMax = max(avg_costs) * 1.1

                chart.bars[0].fillColor = colors.HexColor("#e67e22")
                chart.categoryAxis.labels.boxAnchor = "n"
                chart.categoryAxis.labels.angle = 45

                drawing.add(chart)
                return drawing

        except Exception as e:
            logger.warning(f"Failed to create cost chart: {e}")

        return None

    def _create_recommendations_section(self) -> List:
        """Create recommendations section."""
        elements = []

        elements.append(Paragraph("Recommendations", self.styles["SectionHeader"]))
        elements.append(Spacer(1, 12))

        optimal = self._find_optimal_configuration()
        if optimal:
            recommendations_text = f"""
            <b>Immediate Actions:</b><br/>
            <br/>
            1. <b>Update Memory Configuration:</b> Set your Lambda function memory to {optimal['memory_mb']}MB 
               for optimal performance and cost efficiency.<br/>
            <br/>
            2. <b>Monitor Performance:</b> Implement CloudWatch monitoring to track function performance 
               with the new configuration.<br/>
            <br/>
            3. <b>Cost Tracking:</b> Monitor cost changes after implementing the optimized configuration.<br/>
            <br/>
            <b>Long-term Optimization:</b><br/>
            <br/>
            • Regular performance reviews and re-tuning as workload patterns change<br/>
            • Consider provisioned concurrency for consistent low-latency requirements<br/>
            • Implement automated performance testing as part of CI/CD pipeline<br/>
            • Monitor cold start rates and implement warming strategies if needed<br/>
            """

            elements.append(Paragraph(recommendations_text, self.styles["Normal"]))

        return elements

    def _find_optimal_configuration(self) -> Optional[Dict[str, Any]]:
        """Find optimal configuration based on balanced performance and cost."""
        if not self.configurations:
            return None

        scored_configs = []
        for config in self.configurations:
            executions = config.get("executions", [])
            successful = [e for e in executions if not e.get("error")]

            if successful:
                durations = [e["duration"] for e in successful]
                costs = [
                    calculate_cost(config["memory_mb"], e.get("billed_duration", e["duration"]))
                    for e in successful
                ]

                avg_duration = sum(durations) / len(durations)
                avg_cost = sum(costs) / len(costs)

                # Balanced score (normalize and combine)
                duration_score = 1 / (1 + avg_duration / 1000)
                cost_score = 1 / (1 + avg_cost * 1000000)
                score = duration_score + cost_score

                scored_configs.append(
                    {
                        "memory_mb": config["memory_mb"],
                        "avg_duration": avg_duration,
                        "avg_cost": avg_cost,
                        "score": score,
                    }
                )

        return max(scored_configs, key=lambda x: x["score"]) if scored_configs else None

    def _get_baseline_stats(self) -> Optional[Dict[str, Any]]:
        """Get baseline statistics from first configuration."""
        if not self.configurations:
            return None

        first_config = self.configurations[0]
        executions = first_config.get("executions", [])
        successful = [e for e in executions if not e.get("error")]

        if successful:
            durations = [e["duration"] for e in successful]
            costs = [
                calculate_cost(first_config["memory_mb"], e.get("billed_duration", e["duration"]))
                for e in successful
            ]

            return {
                "memory_mb": first_config["memory_mb"],
                "avg_duration": sum(durations) / len(durations),
                "avg_cost": sum(costs) / len(costs),
            }

        return None

    def _count_total_executions(self) -> int:
        """Count total executions across all configurations."""
        return sum(len(config.get("executions", [])) for config in self.configurations)


class ExcelExporter:
    """Exports Lambda tuning reports to Excel format."""

    def __init__(self, results: Dict[str, Any]):
        """
        Initialize Excel exporter.

        Args:
            results: Tuning results dictionary
        """
        if not EXCEL_AVAILABLE:
            raise ReportGenerationError(
                "pandas and openpyxl are required for Excel export. Install with: pip install pandas openpyxl"
            )

        self.results = results
        self.configurations = results.get("configurations", [])

    def export_comprehensive_report(self, output_path: str, workload_type: str = "generic"):
        """
        Export a comprehensive report with multiple sheets.

        Args:
            output_path: Path to save the Excel file
            workload_type: Type of workload for specialized reporting
        """
        try:
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                # Summary sheet
                self._create_summary_sheet(writer, workload_type)

                # Detailed results sheet
                self._create_detailed_results_sheet(writer)

                # Raw data sheet
                self._create_raw_data_sheet(writer)

                # Cost analysis sheet
                self._create_cost_analysis_sheet(writer)

                # Charts sheet
                self._create_charts_sheet(writer)

            # Add styling and charts
            self._apply_excel_styling(output_path)

            logger.info(f"Excel report exported to {output_path}")

        except Exception as e:
            logger.error(f"Error exporting Excel report: {e}")
            raise ReportGenerationError(f"Failed to export Excel report: {e}")

    def _create_summary_sheet(self, writer: pd.ExcelWriter, workload_type: str):
        """Create summary sheet."""
        optimal = self._find_optimal_configuration()
        baseline = self._get_baseline_stats()

        summary_data = {
            "Metric": [
                "Workload Type",
                "Optimal Memory (MB)",
                "Optimal Duration (ms)",
                "Optimal Cost ($)",
                "Performance Improvement (%)",
                "Cost Savings (%)",
                "Total Configurations",
                "Total Executions",
                "Report Generated",
            ],
            "Value": [
                workload_type.title(),
                optimal["memory_mb"] if optimal else "N/A",
                f"{optimal['avg_duration']:.2f}" if optimal else "N/A",
                f"{optimal['avg_cost']:.6f}" if optimal else "N/A",
                (
                    f"{((baseline['avg_duration'] - optimal['avg_duration']) / baseline['avg_duration'] * 100):.1f}"
                    if optimal and baseline
                    else "N/A"
                ),
                (
                    f"{((baseline['avg_cost'] - optimal['avg_cost']) / baseline['avg_cost'] * 100):.1f}"
                    if optimal and baseline
                    else "N/A"
                ),
                len(self.configurations),
                self._count_total_executions(),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ],
        }

        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

    def _create_detailed_results_sheet(self, writer: pd.ExcelWriter):
        """Create detailed results sheet."""
        results_data = []

        for config in self.configurations:
            memory_mb = config["memory_mb"]
            executions = config.get("executions", [])
            successful = [e for e in executions if not e.get("error")]
            failed = [e for e in executions if e.get("error")]

            if successful:
                durations = [e["duration"] for e in successful]
                costs = [
                    calculate_cost(memory_mb, e.get("billed_duration", e["duration"]))
                    for e in successful
                ]

                stats = calculate_statistics(durations)
                cold_starts = sum(1 for e in successful if e.get("cold_start", False))

                results_data.append(
                    {
                        "Memory (MB)": memory_mb,
                        "Total Executions": len(executions),
                        "Successful": len(successful),
                        "Failed": len(failed),
                        "Success Rate (%)": len(successful) / len(executions) * 100,
                        "Avg Duration (ms)": stats["mean"],
                        "Min Duration (ms)": stats["min"],
                        "Max Duration (ms)": stats["max"],
                        "P95 Duration (ms)": stats.get("p95", 0),
                        "Std Dev Duration (ms)": stats["std"],
                        "Avg Cost ($)": sum(costs) / len(costs),
                        "Total Cost ($)": sum(costs),
                        "Cold Starts": cold_starts,
                        "Cold Start Rate (%)": cold_starts / len(successful) * 100,
                    }
                )

        results_df = pd.DataFrame(results_data)
        results_df.to_excel(writer, sheet_name="Detailed Results", index=False)

    def _create_raw_data_sheet(self, writer: pd.ExcelWriter):
        """Create raw execution data sheet."""
        raw_data = []

        for config in self.configurations:
            memory_mb = config["memory_mb"]
            executions = config.get("executions", [])

            for i, execution in enumerate(executions):
                raw_data.append(
                    {
                        "Memory (MB)": memory_mb,
                        "Execution #": i + 1,
                        "Duration (ms)": execution.get("duration", 0),
                        "Billed Duration (ms)": execution.get(
                            "billed_duration", execution.get("duration", 0)
                        ),
                        "Cost ($)": calculate_cost(
                            memory_mb,
                            execution.get("billed_duration", execution.get("duration", 0)),
                        ),
                        "Cold Start": execution.get("cold_start", False),
                        "Error": execution.get("error", ""),
                        "Timestamp": execution.get("timestamp", ""),
                        "Success": not bool(execution.get("error")),
                    }
                )

        raw_df = pd.DataFrame(raw_data)
        raw_df.to_excel(writer, sheet_name="Raw Data", index=False)

    def _create_cost_analysis_sheet(self, writer: pd.ExcelWriter):
        """Create cost analysis sheet."""
        cost_data = []

        # Monthly projections for different scenarios
        scenarios = [
            {"name": "Low Volume", "invocations_per_month": 100000},
            {"name": "Medium Volume", "invocations_per_month": 1000000},
            {"name": "High Volume", "invocations_per_month": 10000000},
        ]

        for config in self.configurations:
            memory_mb = config["memory_mb"]
            executions = config.get("executions", [])
            successful = [e for e in executions if not e.get("error")]

            if successful:
                costs = [
                    calculate_cost(memory_mb, e.get("billed_duration", e["duration"]))
                    for e in successful
                ]
                avg_cost = sum(costs) / len(costs)

                row = {"Memory (MB)": memory_mb, "Avg Cost per Invocation ($)": avg_cost}

                for scenario in scenarios:
                    monthly_cost = avg_cost * scenario["invocations_per_month"]
                    row[f"{scenario['name']} Monthly ($)"] = monthly_cost
                    row[f"{scenario['name']} Yearly ($)"] = monthly_cost * 12

                cost_data.append(row)

        cost_df = pd.DataFrame(cost_data)
        cost_df.to_excel(writer, sheet_name="Cost Projections", index=False)

    def _create_charts_sheet(self, writer: pd.ExcelWriter):
        """Create charts overview sheet."""
        # Create summary for charts
        chart_data = []

        for config in self.configurations:
            memory_mb = config["memory_mb"]
            executions = config.get("executions", [])
            successful = [e for e in executions if not e.get("error")]

            if successful:
                durations = [e["duration"] for e in successful]
                costs = [
                    calculate_cost(memory_mb, e.get("billed_duration", e["duration"]))
                    for e in successful
                ]

                chart_data.append(
                    {
                        "Memory (MB)": memory_mb,
                        "Avg Duration (ms)": sum(durations) / len(durations),
                        "Avg Cost ($)": sum(costs) / len(costs),
                        "Success Rate (%)": len(successful) / len(executions) * 100,
                    }
                )

        chart_df = pd.DataFrame(chart_data)
        chart_df.to_excel(writer, sheet_name="Chart Data", index=False)

    def _apply_excel_styling(self, output_path: str):
        """Apply styling and charts to Excel file."""
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Re-load the file we just created
            from openpyxl import load_workbook

            wb = load_workbook(output_path)

            # Style the summary sheet
            if "Summary" in wb.sheetnames:
                ws = wb["Summary"]

                # Header styling
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(
                    start_color="3498DB", end_color="3498DB", fill_type="solid"
                )

                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Add chart to Chart Data sheet
            if "Chart Data" in wb.sheetnames:
                ws = wb["Chart Data"]

                # Create bar chart for performance
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.title = "Performance by Memory Configuration"
                chart.y_axis.title = "Duration (ms)"
                chart.x_axis.title = "Memory (MB)"

                # Assume data starts at A1 and we have headers
                data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row, max_col=2)
                cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)

                ws.add_chart(chart, "E2")

            wb.save(output_path)

        except Exception as e:
            logger.warning(f"Failed to apply Excel styling: {e}")

    def _find_optimal_configuration(self) -> Optional[Dict[str, Any]]:
        """Find optimal configuration."""
        if not self.configurations:
            return None

        scored_configs = []
        for config in self.configurations:
            executions = config.get("executions", [])
            successful = [e for e in executions if not e.get("error")]

            if successful:
                durations = [e["duration"] for e in successful]
                costs = [
                    calculate_cost(config["memory_mb"], e.get("billed_duration", e["duration"]))
                    for e in successful
                ]

                avg_duration = sum(durations) / len(durations)
                avg_cost = sum(costs) / len(costs)

                # Balanced score
                duration_score = 1 / (1 + avg_duration / 1000)
                cost_score = 1 / (1 + avg_cost * 1000000)
                score = duration_score + cost_score

                scored_configs.append(
                    {
                        "memory_mb": config["memory_mb"],
                        "avg_duration": avg_duration,
                        "avg_cost": avg_cost,
                        "score": score,
                    }
                )

        return max(scored_configs, key=lambda x: x["score"]) if scored_configs else None

    def _get_baseline_stats(self) -> Optional[Dict[str, Any]]:
        """Get baseline statistics."""
        if not self.configurations:
            return None

        first_config = self.configurations[0]
        executions = first_config.get("executions", [])
        successful = [e for e in executions if not e.get("error")]

        if successful:
            durations = [e["duration"] for e in successful]
            costs = [
                calculate_cost(first_config["memory_mb"], e.get("billed_duration", e["duration"]))
                for e in successful
            ]

            return {
                "memory_mb": first_config["memory_mb"],
                "avg_duration": sum(durations) / len(durations),
                "avg_cost": sum(costs) / len(costs),
            }

        return None

    def _count_total_executions(self) -> int:
        """Count total executions."""
        return sum(len(config.get("executions", [])) for config in self.configurations)


class MultiFormatExporter:
    """Unified exporter supporting multiple formats."""

    def __init__(self, results: Dict[str, Any]):
        """
        Initialize multi-format exporter.

        Args:
            results: Tuning results dictionary
        """
        self.results = results

    def export_report(
        self, output_path: str, format_type: str = "auto", workload_type: str = "generic", **kwargs
    ):
        """
        Export report in specified format.

        Args:
            output_path: Path to save the report
            format_type: Format type ("pdf", "excel", "json", "csv", "auto")
            workload_type: Type of workload
            **kwargs: Additional format-specific options
        """
        if format_type == "auto":
            format_type = self._detect_format_from_path(output_path)

        if format_type == "pdf":
            self._export_pdf(output_path, workload_type, **kwargs)
        elif format_type == "excel":
            self._export_excel(output_path, workload_type, **kwargs)
        elif format_type == "json":
            self._export_json(output_path, **kwargs)
        elif format_type == "csv":
            self._export_csv(output_path, **kwargs)
        else:
            raise ReportGenerationError(f"Unsupported format: {format_type}")

    def _detect_format_from_path(self, output_path: str) -> str:
        """Detect format from file extension."""
        path = Path(output_path)
        extension = path.suffix.lower()

        format_map = {
            ".pdf": "pdf",
            ".xlsx": "excel",
            ".xls": "excel",
            ".json": "json",
            ".csv": "csv",
        }

        return format_map.get(extension, "json")

    def _export_pdf(self, output_path: str, workload_type: str, **kwargs):
        """Export to PDF format."""
        exporter = PDFExporter(self.results)
        exporter.export_performance_report(output_path, workload_type)

    def _export_excel(self, output_path: str, workload_type: str, **kwargs):
        """Export to Excel format."""
        exporter = ExcelExporter(self.results)
        exporter.export_comprehensive_report(output_path, workload_type)

    def _export_json(self, output_path: str, **kwargs):
        """Export to JSON format."""
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        # Enhanced JSON export with metadata
        export_data = {
            "metadata": {
                "export_format": "json",
                "exported_at": datetime.now().isoformat(),
                "tuner_version": "2.0",
                "total_configurations": len(self.results.get("configurations", [])),
                "total_executions": sum(
                    len(c.get("executions", [])) for c in self.results.get("configurations", [])
                ),
            },
            "results": self.results,
            "analysis": {
                "optimal_configuration": self._find_optimal_configuration(),
                "performance_summary": self._generate_performance_summary(),
            },
        }

        with open(output_path, "w") as f:
            json.dump(export_data, f, indent=2, default=str)

        logger.info(f"JSON report exported to {output_path}")

    def _export_csv(self, output_path: str, **kwargs):
        """Export to CSV format."""
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        # Flatten data for CSV
        csv_data = []

        for config in self.results.get("configurations", []):
            memory_mb = config["memory_mb"]
            executions = config.get("executions", [])

            for i, execution in enumerate(executions):
                csv_data.append(
                    {
                        "memory_mb": memory_mb,
                        "execution_number": i + 1,
                        "duration_ms": execution.get("duration", 0),
                        "billed_duration_ms": execution.get(
                            "billed_duration", execution.get("duration", 0)
                        ),
                        "cost_usd": calculate_cost(
                            memory_mb,
                            execution.get("billed_duration", execution.get("duration", 0)),
                        ),
                        "cold_start": execution.get("cold_start", False),
                        "error": execution.get("error", ""),
                        "success": not bool(execution.get("error")),
                        "timestamp": execution.get("timestamp", ""),
                    }
                )

        if EXCEL_AVAILABLE:
            df = pd.DataFrame(csv_data)
            df.to_csv(output_path, index=False)
        else:
            # Fallback to manual CSV writing
            import csv

            if csv_data:
                with open(output_path, "w", newline="") as f:
                    writer = csv.DictWriter(f, fieldnames=csv_data[0].keys())
                    writer.writeheader()
                    writer.writerows(csv_data)

        logger.info(f"CSV report exported to {output_path}")

    def _find_optimal_configuration(self) -> Optional[Dict[str, Any]]:
        """Find optimal configuration."""
        # Reuse logic from other exporters
        pdf_exporter = PDFExporter(self.results)
        return pdf_exporter._find_optimal_configuration()

    def _generate_performance_summary(self) -> Dict[str, Any]:
        """Generate performance summary."""
        optimal = self._find_optimal_configuration()

        if not optimal:
            return {}

        return {
            "optimal_memory_mb": optimal["memory_mb"],
            "optimal_avg_duration_ms": optimal["avg_duration"],
            "optimal_avg_cost_usd": optimal["avg_cost"],
            "total_configurations_tested": len(self.results.get("configurations", [])),
            "total_executions": sum(
                len(c.get("executions", [])) for c in self.results.get("configurations", [])
            ),
        }


# Convenience functions
def export_pdf_report(results: Dict[str, Any], output_path: str, workload_type: str = "generic"):
    """Export PDF report."""
    exporter = PDFExporter(results)
    exporter.export_performance_report(output_path, workload_type)


def export_excel_report(results: Dict[str, Any], output_path: str, workload_type: str = "generic"):
    """Export Excel report."""
    exporter = ExcelExporter(results)
    exporter.export_comprehensive_report(output_path, workload_type)


def export_multi_format(
    results: Dict[str, Any],
    output_path: str,
    format_type: str = "auto",
    workload_type: str = "generic",
):
    """Export report in multiple formats."""
    exporter = MultiFormatExporter(results)
    exporter.export_report(output_path, format_type, workload_type)
